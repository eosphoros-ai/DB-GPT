import io
import json
import logging
import os
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from dbgpt.util.benchmarks.ExcelUtils import ExcelUtils
from dbgpt_serve.evaluate.db.benchmark_db import BenchmarkResultDao

from .models import (
    AnswerExecuteModel,
    BaseInputModel,
    BenchmarkDataSets,
    BenchmarkExecuteConfig,
    DataCompareStrategyConfig,
    OutputType,
    RoundAnswerConfirmModel,
)

logger = logging.getLogger(__name__)


class FileParseService(ABC):
    def __init__(self):
        self._benchmark_dao = BenchmarkResultDao()

        # export column configuration file path
        self._column_config_file_path = os.path.join(
            os.path.dirname(__file__),
            "template",
            "benchmark_column_config_template.json",
        )

    @abstractmethod
    def parse_input_sets(self, path: str) -> BenchmarkDataSets:
        """
        Parse input sets from file
        Args:
            location: File location path
        Returns:
            BenchmarkDataSets: Parsed data sets
        """

    def parse_llm_outputs(self, path: str) -> List[AnswerExecuteModel]:
        data = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                obj = json.loads(line)
                data.append(AnswerExecuteModel.from_dict(obj))
        return data

    @abstractmethod
    def write_data_compare_result(
        self,
        path: str,
        round_id: int,
        confirm_models: List[RoundAnswerConfirmModel],
        is_execute: bool,
        llm_count: int,
    ):
        """Write compare results to File

        Args:
            path: Output file path
            round_id: Round ID
            confirm_models: List of answer confirm models
            is_execute: Whether to execute the comparison
            llm_count: LLM count
        """

    def summary_and_write_multi_round_benchmark_result(
        self, output_path: str, round_id: int
    ) -> str:
        """Compute summary from the Excel file grouped by llmCode and return JSON list.

        It reads the '<base>_round{round_id}.xlsx' file and sheet
        'benchmark_compare_result', then for each llmCode counts the compareResult
         column (RIGHT/WRONG/FAILED/EXCEPTION) to build summary list.
        """
        try:
            base_name = Path(output_path).stem
            extension = Path(output_path).suffix
            if extension.lower() not in [".xlsx", ".xls"]:
                extension = ".xlsx"
            excel_file = Path(output_path).parent / f"{base_name}{extension}"
            if not excel_file.exists():
                logger.warning(f"summary excel not found: {excel_file}")
                return json.dumps([], ensure_ascii=False)

            df = pd.read_excel(str(excel_file), sheet_name="benchmark_compare_result")
            if "compareResult" not in df.columns:
                logger.warning("compareResult column missing in excel")
                return json.dumps([], ensure_ascii=False)

            # ensure llmCode column exists
            if "llmCode" not in df.columns:
                df["llmCode"] = None

            summaries = []
            for llm_code, group in df.groupby("llmCode"):
                right = int((group["compareResult"] == "RIGHT").sum())
                wrong = int((group["compareResult"] == "WRONG").sum())
                failed = int((group["compareResult"] == "FAILED").sum())
                exception = int((group["compareResult"] == "EXCEPTION").sum())
                summaries.append(
                    {
                        "llmCode": None if pd.isna(llm_code) else str(llm_code),
                        "right": right,
                        "wrong": wrong,
                        "failed": failed,
                        "exception": exception,
                    }
                )

            logger.info(
                f"[summary] computed per llmCode for round={round_id},"
                f" output_path={output_path} -> {summaries}"
            )
            return json.dumps(summaries, ensure_ascii=False)
        except Exception as e:
            logger.error(f"summary compute error from excel: {e}", exc_info=True)
            return json.dumps([], ensure_ascii=False)

    def get_input_stream(self, location: str):
        """Get input stream from location

        Args:
            location: File location path

        Returns:
            Optional[io.BytesIO]: File input stream or None if not found
        """
        try:
            with open(location, "rb") as file:
                return io.BytesIO(file.read())
        except FileNotFoundError:
            logger.error(f"file Not found with: {location}")
            return None
        except Exception as e:
            logger.error(f"Error reading file: {location}, errorMsg: {e}")
            return None

    @abstractmethod
    def parse_standard_benchmark_sets(
        self, standard_excel_path: str
    ) -> List[AnswerExecuteModel]:
        """
        Parse standard benchmark sets from file.
        This method must be implemented by subclasses.

        Args:
            standard_excel_path: Path to the standard benchmark file

        Returns:
            List[AnswerExecuteModel]: List of parsed answer execute models
        """
        pass

    @abstractmethod
    def write_multi_round_benchmark_result(
        self,
        output_file_path: str,
        round_id: int,
        config: BenchmarkExecuteConfig,
        inputs: List[BaseInputModel],
        outputs: List[OutputType],
        start_index: int,
        offset: int,
    ) -> bool:
        """
        Write Benchmark Task Multi round Result

        Args:
            output_file_path: Output file path
            round_id: Round ID
            config: Benchmark configuration
            inputs: List of input data
            outputs: List of output data
            start_index: Starting index (batch start row index)
            offset: Offset(file rows offset)
        """


class ExcelFileParseService(FileParseService):
    def parse_input_sets(self, path: str) -> BenchmarkDataSets:
        """
        Parse input sets from excel file
        Args:
            path: File location path
        Returns:
            BenchmarkDataSets: Parsed data sets
        """
        input_stream = self.get_input_stream(path)

        if input_stream is None:
            raise RuntimeError(f"file not found! path: {path}")

        # Parse excel file to get data sets
        input_sets = BenchmarkDataSets()
        workbook = None

        try:
            workbook = load_workbook(input_stream, data_only=True)
            input_list = []

            # Get the first worksheet
            sheet = workbook.worksheets[0]

            for row_num in range(
                2, sheet.max_row + 1
            ):  # Skip header row (start from 1-based index)
                row = sheet[row_num]
                if ExcelUtils.is_row_empty(row):
                    continue

                # Get content from columns 1-6 (0-based index 0-5)
                serial_no_cell = row[0]
                analysis_model_id_cell = row[1]
                question_cell = row[2]
                self_define_tags_cell = row[3]
                knowledge_cell = row[4]
                llm_output_cell = row[5]
                prompt_cell = row[8]

                # Build input model
                input_model = BaseInputModel(
                    serial_no=int(
                        ExcelUtils.get_cell_value_as_string(serial_no_cell) or "0"
                    ),
                    analysis_model_id=ExcelUtils.get_cell_value_as_string(
                        analysis_model_id_cell
                    ),
                    question=ExcelUtils.get_cell_value_as_string(question_cell),
                    self_define_tags=ExcelUtils.get_cell_value_as_string(
                        self_define_tags_cell
                    ),
                    llm_output=ExcelUtils.get_cell_value_as_string(llm_output_cell),
                    knowledge=ExcelUtils.get_cell_value_as_string(knowledge_cell),
                    prompt=ExcelUtils.get_cell_value_as_string(prompt_cell),
                )

                input_list.append(input_model)

            input_sets.data_list = input_list
        except Exception as e:
            logger.error(f"parse excel error, path: {path}, errorMsg: {e}")
        finally:
            try:
                if workbook is not None:
                    workbook.close()
            except Exception as e:
                logger.error(f"close workbook error, path: {path}, errorMsg: {e}")

        return input_sets

    def parse_standard_benchmark_sets(
        self, standard_excel_path: str
    ) -> List[AnswerExecuteModel]:
        df = pd.read_excel(standard_excel_path, sheet_name=0)
        outputs: List[AnswerExecuteModel] = []
        for _, row in df.iterrows():
            try:
                serial_no = int(row["编号"])
            except Exception:
                continue
            question = row.get("用户问题")
            analysis_model_id = row.get("数据集ID")
            llm_output = (
                None if pd.isna(row.get("标准答案SQL")) else str(row.get("标准答案SQL"))
            )
            order_by = True
            if not pd.isna(row.get("是否排序")):
                try:
                    order_by = bool(int(row.get("是否排序")))
                except Exception:
                    order_by = True

            std_result: Optional[List[Dict[str, List[str]]]] = None
            if not pd.isna(row.get("标准结果")):
                std_result_raw = str(row.get("标准结果"))
                std_result = self._parse_multi_standard_result(std_result_raw)

            strategy_config = DataCompareStrategyConfig(
                strategy="CONTAIN_MATCH",
                order_by=order_by,
                standard_result=std_result if std_result is not None else None,
            )
            outputs.append(
                AnswerExecuteModel(
                    serialNo=serial_no,
                    analysisModelId=analysis_model_id,
                    question=question,
                    llmOutput=llm_output,
                    executeResult=std_result,
                    strategyConfig=strategy_config,
                )
            )
        return outputs

    def write_multi_round_benchmark_result(
        self,
        output_file_path: str,
        round_id: int,
        config: BenchmarkExecuteConfig,
        inputs: List[BaseInputModel],
        outputs: List[OutputType],
        start_index: int,
        offset: int,
    ) -> bool:
        """
        Write the benchmark Result to Excel File With Multi Round

        Args:
            output_file_path: Output file path
            round_id: Round ID
            config: Benchmark configuration
            inputs: List of input data
            outputs: List of output data
            start_index: Starting index (batch start row index)
            offset: Offset(file rows offset)

        Returns:
            bool: Returns True if write is successful, False otherwise
        """
        try:
            # 确保输出目录存在
            output_dir = Path(output_file_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 从JSON文件加载列配置
            column_config = self._load_column_config()

            # 按index排序确保列顺序正确
            column_config.sort(key=lambda x: x["index"])

            # 创建表头
            headers = [col["header"] for col in column_config]

            base_name = Path(output_file_path).stem
            extension = Path(output_file_path).suffix
            if extension.lower() not in [".xlsx", ".xls"]:
                extension = ".xlsx"

            output_file = output_dir / f"{base_name}{extension}"

            # 创建输入数据映射，便于查找
            input_map = {inp.serial_no: inp for inp in inputs}

            # 准备数据行
            data_rows = []

            for output in outputs:
                # 获取对应的输入数据
                input_data = input_map.get(output.serialNo)

                # 构建数据行
                row_data = []
                for col in column_config:
                    field = col["field"]
                    source_type = col["sourceType"]
                    processor_type = col["processorType"]

                    # Determine data source based on sourceType
                    value = self._get_value_by_source_type(
                        field, source_type, processor_type, input_data, output, round_id
                    )
                    # Process value based on processorType
                    value = self._process_value_by_type(value, processor_type)

                    row_data.append(value)
                data_rows.append(row_data)

            # 检查文件是否存在
            if output_file.exists():
                # 文件存在，读取现有工作簿
                workbook = load_workbook(str(output_file))
                if "dataset_evaluation_result" in workbook.sheetnames:
                    worksheet = workbook["dataset_evaluation_result"]
                else:
                    worksheet = workbook.create_sheet("dataset_evaluation_result")
            else:
                # 文件不存在，创建新工作簿
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "dataset_evaluation_result"

                # 写入表头（第1行）
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)

            # 计算写入的起始行号 公式：start_index + offset + 2
            # (+1是因为Excel行号从1开始，+1是因为表头占一行)
            write_start_row = start_index + offset + 2

            # 写入数据行
            for row_idx, row_data in enumerate(data_rows):
                excel_row = write_start_row + row_idx
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=excel_row, column=col_idx, value=value)

            # 调整列宽以适应内容
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        logger.warning(f"error while compute column length: {str(e)}")
                # 设置列宽，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 保存工作簿
            workbook.save(str(output_file))
            workbook.close()

            logger.info(
                f"write excel file success: {output_file}, "
                f"write_start_row: {write_start_row}, "
                f"data_rows: {len(data_rows)}, "
                f"start_index: {start_index}, offset: {offset}"
            )
            return True

        except Exception as e:
            logger.error(f"write excel file error: {e}", exc_info=True)
            return False

    def write_data_compare_result(
        self,
        path: str,
        round_id: int,
        confirm_models: List[RoundAnswerConfirmModel],
        is_execute: bool,
        llm_count: int,
    ):
        """Write compare results to an Excel file

        The output Excel file will be named as '<base>.xlsx' and
        sheet name is 'benchmark_compare_result'. If the file exists, it will
        append rows; otherwise it will create a new file with headers.
        """
        try:
            # Ensure output directory exists
            output_dir = Path(path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            output_file = path

            headers = [
                "serialNo",
                "analysisModelId",
                "question",
                "selfDefineTags",
                "prompt",
                "standardAnswerSql",
                "standardAnswer",
                "llmCode",
                "llmOutput",
                "executeResult",
                "errorMsg",
                "compareResult",
            ]

            # Load or create workbook and sheet
            if Path(output_file).exists():
                workbook = load_workbook(str(output_file))
                if "benchmark_compare_result" in workbook.sheetnames:
                    worksheet = workbook["benchmark_compare_result"]
                else:
                    worksheet = workbook.create_sheet("benchmark_compare_result")
                    # Write headers if new sheet
                    for col_idx, header in enumerate(headers, 1):
                        worksheet.cell(row=1, column=col_idx, value=header)
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "benchmark_compare_result"
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)

            # Determine start row to append
            start_row = worksheet.max_row + 1 if worksheet.max_row else 2

            # Append rows
            for idx, cm in enumerate(confirm_models):
                row_data = [
                    cm.serialNo,
                    cm.analysisModelId,
                    cm.question,
                    cm.selfDefineTags,
                    cm.prompt,
                    cm.standardAnswerSql,
                    self._format_set_result(cm.strategyConfig.standard_result)
                    if cm.strategyConfig is not None
                    else "",
                    cm.llmCode,
                    cm.llmOutput,
                    json.dumps(cm.executeResult, ensure_ascii=False)
                    if cm.executeResult is not None
                    else "",
                    cm.errorMsg,
                    cm.compareResult.value if cm.compareResult else None,
                ]
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=start_row + idx, column=col_idx, value=value)

            # Autosize columns (simple strategy)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 80)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(str(output_file))
            workbook.close()
            logger.info(
                f"[write_data_compare_result] compare written to Excel: {output_file}"
            )
        except Exception as e:
            logger.error(
                f"[write_data_compare_result] write excel error for path={path}: {e}"
            )

    def _get_value_by_source_type(
        self,
        field: str,
        source_type: str,
        processor_type: str,
        input_data,
        output,
        round_id: int,
    ) -> Any:
        """
        Get the value based on the source type

        Args:
            field: Field name
            source_type: Source type
            processor_type: Processor type
            input_data: Input data
            output: Output data
            round_id: Round ID

        Returns:
            The retrieved value
        """
        value = None

        # 根据sourceType确定数据来源
        if source_type == "INPUT" and input_data:
            # 从输入数据获取
            if field == "serialNo":
                value = input_data.serial_no
            elif field == "llmCode":
                value = getattr(input_data, "llm_code", "")
            elif field == "analysisModelId":
                value = input_data.analysis_model_id
            elif field == "question":
                value = input_data.question
            elif field == "selfDefineTags":
                value = input_data.self_define_tags
            elif field == "knowledge":
                value = input_data.knowledge
            elif field == "prompt":
                value = input_data.prompt
        elif source_type == "PARAM":
            # 从参数获取
            if field == "roundId":
                value = str(round_id)
        elif source_type == "OUTPUT":
            # 从输出数据获取
            if field == "cotLength":
                value = getattr(output, "cotTokens", 0) or 0
            elif field == "llmOutput":
                value = output.llmOutput
            elif field == "executeResult":
                # JSON处理器：将字典转换为JSON字符串
                if processor_type == "JsonProcessor":
                    value = (
                        json.dumps(output.executeResult, ensure_ascii=False)
                        if output.executeResult
                        else ""
                    )
                else:
                    value = str(output.executeResult) if output.executeResult else ""
            elif field == "errorMsg":
                value = output.errorMsg
            elif field == "traceId":
                value = ""  # traceId需要从其他地方获取，这里暂时留空
            elif field == "costTime":
                value = getattr(output, "cost_time", "") or ""

        return value

    def _process_value_by_type(self, value, processor_type: str) -> Any:
        """
        Process value based on processor type
        Args:
            value: Original value
            processor_type: Processor type
        """
        if processor_type == "IntegerProcessor":
            try:
                return int(value) if value is not None else 0
            except (ValueError, TypeError):
                return 0
        elif processor_type == "LongProcessor":
            try:
                return int(value) if value is not None else 0
            except (ValueError, TypeError):
                return 0
        elif processor_type in [
            "StringProcessor",
            "LongTextProcessor",
            "JsonProcessor",
        ]:
            return str(value) if value is not None else ""
        else:
            return str(value) if value is not None else ""

    def _parse_multi_standard_result(
        self, std_result_raw: str
    ) -> Optional[List[Dict[str, List[str]]]]:
        """
        Parse multiple standard results from raw string data.

        Handles multiple results separated by newlines and parses each line as a dict.

        Args:
            std_result_raw (str): Raw standard result string with multiple lines

        Returns:
            Optional[List[Dict[str, List[str]]]]: List of parsed dictionaries,
            or None if parsing fails or no valid data
        """
        try:
            std_result_raw = std_result_raw.strip()
            if not std_result_raw:
                return None

            # 处理多个结果，通过换行符分隔
            result_lines = std_result_raw.split("\n")
            result_list = []

            for line in result_lines:
                line = line.strip()
                if line:
                    try:
                        result_list.append(json.loads(line))
                    except Exception as e:
                        logger.warning(
                            f"Failed to parse line as JSON: {line}, error: {e}"
                        )
                        continue

            return result_list if result_list else None
        except Exception as e:
            logger.error(f"parse multiple standard results error: {e}")
            return None

    def _format_set_result(
        self, sql_result: List[Dict[str, List[str]]]
    ) -> Optional[str]:
        """
        Format Multi StandardAnswer result
        Returns:
            Optional[str]: Formatted result string with newline separators
        """
        if not sql_result:
            return None

        result_list = []
        for result in sql_result:
            result_list.append(json.dumps(result, ensure_ascii=False))

        return "\n".join(result_list)

    def _load_column_config(self) -> List[Dict]:
        """
        Load column configuration from JSON file

        Returns:
            List[Dict]: List of column configurations
        """
        try:
            with open(self._column_config_file_path, "r", encoding="utf-8") as file:
                config_data = json.load(file)
                return config_data.get("columns", [])
        except Exception as e:
            logger.error(
                f"Failed to load column configuration file: {e},"
                f" using default configuration"
            )
            raise ValueError("Failed to load column configuration file")
