import json
import logging
import os
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Dict, List, Optional, Union

from dbgpt.component import ComponentType, SystemApp
from dbgpt.core import LLMClient
from dbgpt.model import DefaultLLMClient
from dbgpt.model.cluster import WorkerManagerFactory
from dbgpt.storage.metadata import BaseDao
from dbgpt.util.benchmarks import StorageUtil
from dbgpt.util import get_or_create_event_loop

from ....core import BaseService
from ....prompt.service.service import Service as PromptService
from ....rag.service.service import Service as RagService
from ....rag.storage_manager import StorageManager
from ...api.schemas import EvaluateServeRequest, EvaluateServeResponse
from ...config import ServeConfig
from ...models.models import ServeDao, ServeEntity
from .benchmark_llm_task import BenchmarkLLMTask
from .data_compare_service import DataCompareService
from .file_parse_service import ExcelFileParseService, FileParseService
from .models import (
    BaseInputModel,
    BenchmarkDataSets,
    BenchmarkExecuteConfig,
    BenchmarkModeTypeEnum,
    BenchmarkTaskResult,
    ContentTypeEnum,
    FileParseTypeEnum,
    InputType,
    OutputType,
    ReasoningResponse,
)
from .user_input_execute_service import UserInputExecuteService

logger = logging.getLogger(__name__)

executor = ThreadPoolExecutor(max_workers=5)

BENCHMARK_SERVICE_COMPONENT_NAME = "dbgpt_serve_evaluate_benchmark_service"

STANDARD_BENCHMARK_FILE_PATH = (
    "pilot/benchmark_meta_data/"
    "2025_07_27_public_500_standard_benchmark_question_list_v2.xlsx"
)


def get_rag_service(system_app) -> RagService:
    return system_app.get_component("dbgpt_rag_service", RagService)


def get_prompt_service(system_app) -> PromptService:
    return system_app.get_component("dbgpt_serve_prompt_service", PromptService)


class BenchmarkService(
    BaseService[ServeEntity, EvaluateServeRequest, EvaluateServeResponse]
):
    """The benchmark service class for Evaluate"""

    name = BENCHMARK_SERVICE_COMPONENT_NAME
    batch_lock = threading.RLock()

    def __init__(
        self, system_app: SystemApp, config: ServeConfig, dao: Optional[ServeDao] = None
    ):
        self._system_app = system_app
        self._serve_config: ServeConfig = config
        self._dao: ServeDao = dao
        super().__init__(system_app)
        self.rag_service = get_rag_service(system_app)
        self.prompt_service = get_prompt_service(system_app)

        fps = FileParseService()
        dcs = DataCompareService()
        self.user_input_execute_service = UserInputExecuteService(fps, dcs)

        self.trigger_executor = ThreadPoolExecutor(
            max_workers=5, thread_name_prefix="benchmark-fileWrite"
        )
        
        # 设置列配置文件路径
        self._column_config_file_path = os.path.join(
            os.path.dirname(__file__), 
            "template", 
            "benchmark_column_config_template.json"
        )

    def init_app(self, system_app: SystemApp) -> None:
        """Initialize the service

        Args:
            system_app (SystemApp): The system app
        """
        self._system_app = system_app

    @property
    def storage_manager(self):
        return StorageManager.get_instance(self._system_app)

    @property
    def dao(self) -> BaseDao[ServeEntity, EvaluateServeRequest, EvaluateServeResponse]:
        """Returns the internal DAO."""
        return self._dao

    @property
    def config(self) -> ServeConfig:
        """Returns the internal ServeConfig."""
        return self._serve_config

    @property
    def llm_client(self) -> LLMClient:
        worker_manager = self._system_app.get_component(
            ComponentType.WORKER_MANAGER_FACTORY, WorkerManagerFactory
        ).create()
        return DefaultLLMClient(worker_manager, True)

    def _load_column_config(self) -> List[Dict]:
        """
        Load column configuration from JSON file

        Returns:
            List[Dict]: List of column configurations
        """
        try:
            with open(self._column_config_file_path, 'r', encoding='utf-8') as file:
                config_data = json.load(file)
                return config_data.get("columns", [])
        except Exception as e:
            logger.error(f"Failed to load column configuration file: {e},"
                         f" using default configuration")
            return [
                {
                    "index": 0,
                    "header": "编号",
                    "field": "serialNo",
                    "sourceType": "INPUT",
                    "processorType": "IntegerProcessor",
                },
                {
                    "index": 1,
                    "header": "大模型名称",
                    "field": "llmCode",
                    "sourceType": "INPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 2,
                    "header": "轮次",
                    "field": "roundId",
                    "sourceType": "PARAM",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 3,
                    "header": "数据集ID",
                    "field": "analysisModelId",
                    "sourceType": "INPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 4,
                    "header": "用户问题",
                    "field": "question",
                    "sourceType": "INPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 5,
                    "header": "自定义标签",
                    "field": "selfDefineTags",
                    "sourceType": "INPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 6,
                    "header": "知识",
                    "field": "knowledge",
                    "sourceType": "INPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 7,
                    "header": "prompt",
                    "field": "prompt",
                    "sourceType": "INPUT",
                    "processorType": "LongTextProcessor",
                },
                {
                    "index": 8,
                    "header": "Cot长度",
                    "field": "cotLength",
                    "sourceType": "OUTPUT",
                    "processorType": "LongProcessor",
                },
                {
                    "index": 9,
                    "header": "LLM输出结果",
                    "field": "llmOutput",
                    "sourceType": "OUTPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 10,
                    "header": "结果执行",
                    "field": "executeResult",
                    "sourceType": "OUTPUT",
                    "processorType": "JsonProcessor",
                },
                {
                    "index": 11,
                    "header": "执行结果的报错信息",
                    "field": "errorMsg",
                    "sourceType": "OUTPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 12,
                    "header": "traceId",
                    "field": "traceId",
                    "sourceType": "OUTPUT",
                    "processorType": "StringProcessor",
                },
                {
                    "index": 13,
                    "header": "耗时（秒）",
                    "field": "costTime",
                    "sourceType": "OUTPUT",
                    "processorType": "StringProcessor",
                },
            ]

    async def run_dataset_benchmark(
        self,
        evaluate_code: str,
        scene_key: str,
        scene_value: str,
        input_file_path: str,
        output_file_path: str,
        model_list: List[str],
    ) -> List[BenchmarkTaskResult[OutputType]]:
        """
        Run the dataset benchmark
        """
        logger.info(
            f"Run dataset benchmark, evaluate_code: {evaluate_code},"
            f" scene_key: {scene_key}, scene_value: {scene_value},"
            f" input_file_path: {input_file_path}, output_file_path: {output_file_path}"
        )
        if not input_file_path:
            input_file_path = STANDARD_BENCHMARK_FILE_PATH

        config = await self._build_benchmark_config(model_list, output_file_path)

        # read input file
        input_list: List[BaseInputModel] = self.read_input_file(input_file_path)

        result_list = []
        for i in range(1, config.round_time + 1):
            round_result_list = []

            llm_index = 0
            for llm_code, thread_num in config.llm_thread_map.items():
                # 每个llm_code对应的偏移量：llm_index * input_list长度
                offset = len(input_list) * llm_index

                llm_result = BenchmarkTaskResult[OutputType]()
                try:
                    llm_result = self.batch_execute(
                        config,
                        input_list,
                        llm_code,
                        thread_num,
                        i,
                        output_file_path,
                        offset,
                    )
                except Exception as e:
                    logger.error(f"batchExecute error! {e}")

                if llm_result is not None:
                    round_result_list.append(llm_result)
                    llm_index += 1

            self.user_input_execute_service.post_dispatch(
                i,
                config,
                input_list,
                round_result_list,
                input_file_path,
                output_file_path,
            )
            result_list.extend(round_result_list)

        return result_list

    def read_input_file(
        self, input_file_path: str
    ) -> Union[List[BaseInputModel], None]:
        file_parse_type: FileParseTypeEnum = StorageUtil.get_file_parse_type(
            input_file_path
        )
        if file_parse_type == FileParseTypeEnum.EXCEL:
            input_sets: BenchmarkDataSets = ExcelFileParseService().parse_input_sets(
                input_file_path
            )
            return input_sets.data_list
        return None

    async def _build_benchmark_config(self, model_list, output_file_path):
        config = BenchmarkExecuteConfig(
            benchmark_mode_type=BenchmarkModeTypeEnum.EXECUTE,
            standard_file_path=STANDARD_BENCHMARK_FILE_PATH,
        )
        config.output_file_path = output_file_path
        config.content_type = ContentTypeEnum.SQL
        config.round_time = 1
        config.thread_num = 1
        config.execute_llm_result = True
        config.invoke_llm = True
        config.compare_result_enable = True
        config.file_parse_type = FileParseTypeEnum.EXCEL
        config.llm_thread_map = {model: 1 for model in model_list}
        return config

    def batch_execute(
        self,
        config: BenchmarkExecuteConfig,
        inputs: List[InputType],
        llm_code: str,
        thread_num: int,
        round_id: int,
        output_file_path: str,
        offset: int,
    ) -> BenchmarkTaskResult[OutputType]:
        """
        Batch execute the benchmark Task with LLM
        """
        result = BenchmarkTaskResult[OutputType]()
        result.trace_id = str(uuid.uuid4()).replace("-", "")
        result.task_id = str(uuid.uuid4())
        result.start_time = datetime.now()

        executor = ThreadPoolExecutor(
            max_workers=thread_num, thread_name_prefix="benchmark-USER_INPUT_EXECUTE"
        )

        output_sets = BenchmarkDataSets[OutputType]()
        output_list = []

        written_batches = set()  # 记录已写入批次
        complete_map = {}  # 记录任务完成状态，使用Dict[int, OutputType]

        # 线程锁，保证线程安全
        lock = threading.Lock()

        def execute_task(input_data: InputType):
            """task execution function"""
            try:
                input_data.llm_code = llm_code
                logger.info(
                    f"[benchmark_task]start executeBenchmark!"
                    f" input={json.dumps(input_data.to_dict(), ensure_ascii=False)}"
                )

                start_time = time.time()
                loop = get_or_create_event_loop()
                output: OutputType = loop.run_until_complete(
                    self.execute(config, input_data)
                )
                cost_time = int(time.time() - start_time)

                output.cost_time = cost_time
                logger.info(
                    f"[benchmark_task]end executeBenchmark!"
                    f" output={json.dumps(output.to_dict(), ensure_ascii=False)}"
                )

                # 线程安全地添加结果
                with lock:
                    output_list.append(output)

                # 检查并触发批次处理
                self.check_and_trigger_batch(
                    output,
                    inputs,
                    round_id,
                    config,
                    output_file_path,
                    written_batches,
                    complete_map,
                    offset,
                )
            except Exception as e:
                logger.error(f"executeSingleTimeBenchmark error, error: {e}")

        # 提交所有任务
        futures = [executor.submit(execute_task, input_data) for input_data in inputs]

        # 等待所有任务完成
        for future in futures:
            future.result()

        executor.shutdown(wait=True)

        output_sets.data_list = output_list
        result.benchmark_data_sets = output_sets
        return result

    async def execute(
            self, config: BenchmarkExecuteConfig, input: InputType
    ) -> Union[OutputType, None]:
        """
        Execute LLM Benchmark Task
        """
        try:
            # 1. 组装评测输入
            if input.prompt is None:
                raise Exception("benchmark datasets not have prompt template!")
            input.prompt = input.prompt

            # 2. 执行评测 - 使用同步方式调用异步方法
            benchmark_llm_task_service = BenchmarkLLMTask(
                llm_client=self.llm_client, model_name=input.llm_code
            )

            response: ReasoningResponse = await (
                benchmark_llm_task_service.invoke_llm(prompt=input.prompt)
            )

            # 3. 组装评测输出
            return await self.user_input_execute_service.build_output(config, input, response)
        except Exception as e:
            logger.error(
                f"execute benchmark error!  error: {e}"
            )
        return None

    def check_and_trigger_batch(
        self,
        output: OutputType,
        inputs: List[BaseInputModel],
        round_id: int,
        config: BenchmarkExecuteConfig,
        output_file_path: str,
        written_batches: set,
        complete_map: Dict[int, OutputType],
        offset: int,
    ):
        """
        Check if all tasks in the current batch are completed
        and trigger batch processing.
        """
        with self.batch_lock:
            complete_map[output.serialNo] = output
            batch_size = 10

            # 查找当前任务的索引
            task_index = -1
            for i in range(len(inputs)):
                if inputs[i].serial_no == output.serialNo:
                    task_index = i
                    break

            batch_index = task_index // batch_size

            # 避免重复写入
            if batch_index in written_batches:
                return

            # 计算批次范围
            batch_start = batch_index * batch_size
            batch_end = min(len(inputs), (batch_index + 1) * batch_size)

            # 检查当前批次是否全部完成
            is_batch_complete = True
            batch_outputs: List[OutputType] = []

            for i in range(batch_start, batch_end):
                serial_no = inputs[i].serial_no
                if serial_no not in complete_map:
                    is_batch_complete = False
                    break
                batch_outputs.append(complete_map[serial_no])

            if is_batch_complete:
                # 触发异步任务并写入
                def batch_write_task():
                    try:
                        # 执行写入逻辑
                        batch_outputs.sort(key=lambda x: x.serialNo)
                        self.write_output_file(
                            output_file_path,
                            round_id,
                            config,
                            inputs,
                            batch_outputs,
                            batch_start,
                            offset,
                        )
                    except Exception as e:
                        logger.error(f"Batch write error: {e}")

                self.trigger_executor.submit(batch_write_task)
                written_batches.add(batch_index)

    def write_output_file(
        self,
        output_file_path: str,
        round_id: int,
        config: BenchmarkExecuteConfig,
        inputs: List[BaseInputModel],
        outputs: List[OutputType],
        start_index: int,
        offset: int,
    ) -> bool:
        """
        Write the output file

        Args:
            output_file_path: Output file path
            round_id: Round ID
            config: Benchmark configuration
            inputs: List of input data
            outputs: List of output data
            start_index: Starting index (batch start row index)
            offset: Offset(file rows offset)

        Returns:
            bool: Returns True if write is successful, False otherwise
        """
        try:
            from pathlib import Path
            import pandas as pd
            from openpyxl import load_workbook, Workbook

            # 确保输出目录存在
            output_dir = Path(output_file_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 从JSON文件加载列配置
            column_config = self._load_column_config()

            # 按index排序确保列顺序正确
            column_config.sort(key=lambda x: x["index"])

            # 创建表头
            headers = [col["header"] for col in column_config]

            # 构造文件名：每个round_id一个文件
            base_name = Path(output_file_path).stem
            extension = Path(output_file_path).suffix
            if extension.lower() not in [".xlsx", ".xls"]:
                extension = ".xlsx"
            
            output_file = output_dir / f"{base_name}_round{round_id}{extension}"

            # 创建输入数据映射，便于查找
            input_map = {inp.serial_no: inp for inp in inputs}

            # 准备数据行
            data_rows = []

            for output in outputs:
                # 获取对应的输入数据
                input_data = input_map.get(output.serialNo)

                # 构建数据行
                row_data = []
                for col in column_config:
                    field = col["field"]
                    source_type = col["sourceType"]
                    processor_type = col["processorType"]

                    value = None

                    # 根据sourceType确定数据来源
                    if source_type == "INPUT" and input_data:
                        # 从输入数据获取
                        if field == "serialNo":
                            value = input_data.serial_no
                        elif field == "llmCode":
                            value = getattr(input_data, "llm_code", "")
                        elif field == "analysisModelId":
                            value = input_data.analysis_model_id
                        elif field == "question":
                            value = input_data.question
                        elif field == "selfDefineTags":
                            value = input_data.self_define_tags
                        elif field == "knowledge":
                            value = input_data.knowledge
                        elif field == "prompt":
                            value = input_data.prompt
                    elif source_type == "PARAM":
                        # 从参数获取
                        if field == "roundId":
                            value = str(round_id)
                    elif source_type == "OUTPUT":
                        # 从输出数据获取
                        if field == "cotLength":
                            value = getattr(output, "cotTokens", 0) or 0
                        elif field == "llmOutput":
                            value = output.llmOutput
                        elif field == "executeResult":
                            # JSON处理器：将字典转换为JSON字符串
                            if processor_type == "JsonProcessor":
                                value = (
                                    json.dumps(output.executeResult, ensure_ascii=False)
                                    if output.executeResult
                                    else ""
                                )
                            else:
                                value = (
                                    str(output.executeResult)
                                    if output.executeResult
                                    else ""
                                )
                        elif field == "errorMsg":
                            value = output.errorMsg
                        elif field == "traceId":
                            value = ""  # traceId需要从其他地方获取，这里暂时留空
                        elif field == "costTime":
                            value = getattr(output, "cost_time", "") or ""

                    # 根据processorType处理值
                    if processor_type == "IntegerProcessor":
                        try:
                            value = int(value) if value is not None else 0
                        except (ValueError, TypeError):
                            value = 0
                    elif processor_type == "LongProcessor":
                        try:
                            value = int(value) if value is not None else 0
                        except (ValueError, TypeError):
                            value = 0
                    elif processor_type in [
                        "StringProcessor",
                        "LongTextProcessor",
                        "JsonProcessor",
                    ]:
                        value = str(value) if value is not None else ""
                    else:
                        value = str(value) if value is not None else ""

                    row_data.append(value)

                data_rows.append(row_data)

            # 检查文件是否存在
            if output_file.exists():
                # 文件存在，读取现有工作簿
                workbook = load_workbook(str(output_file))
                if "dataset_evaluation_result" in workbook.sheetnames:
                    worksheet = workbook["dataset_evaluation_result"]
                else:
                    worksheet = workbook.create_sheet("dataset_evaluation_result")
            else:
                # 文件不存在，创建新工作簿
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "dataset_evaluation_result"
                
                # 写入表头（第1行）
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)

            # 计算写入的起始行号
            # 公式：start_index + offset + 2（+1是因为Excel行号从1开始，+1是因为表头占一行）
            write_start_row = start_index + offset + 2

            # 写入数据行
            for row_idx, row_data in enumerate(data_rows):
                excel_row = write_start_row + row_idx
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=excel_row, column=col_idx, value=value)

            # 调整列宽以适应内容
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        logger.warning(
                            f"error while compute column length: {str(e)}"
                        )
                # 设置列宽，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 保存工作簿
            workbook.save(str(output_file))
            workbook.close()

            logger.info(
                f"write excel file success: {output_file}, "
                f"write_start_row: {write_start_row}, "
                f"data_rows: {len(data_rows)}, "
                f"start_index: {start_index}, offset: {offset}"
            )
            return True

        except Exception as e:
            logger.error(f"write excel file error: {e}", exc_info=True)
            return False
