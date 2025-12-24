import json
import logging
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from dbgpt.util.benchmarks.ExcelUtils import ExcelUtils

from ..file_parse_service import FileParseService
from ..models import (
    AnswerExecuteModel,
    BaseInputModel,
    BenchmarkDataSets,
    DataCompareStrategyConfig, EvaluationEnv,
)

logger = logging.getLogger(__name__)


class ExcelFileParseService(FileParseService):
    def parse_input_sets(self, path: str, evaluation_env: EvaluationEnv) -> BenchmarkDataSets:
        """
        Parse input sets from excel file
        Args:
            path: File location path
            evaluation_env: Evaluation environment
        Returns:
            BenchmarkDataSets: Parsed data sets
        """
        input_stream = self.get_input_stream(path)

        if input_stream is None:
            raise RuntimeError(f"file not found! path: {path}")

        # Parse excel file to get data sets
        input_sets = BenchmarkDataSets()
        workbook = None

        try:
            workbook = load_workbook(input_stream, data_only=True)
            input_list = []

            # Get the first worksheet
            sheet = workbook.worksheets[0]

            for row_num in range(
                2, sheet.max_row + 1
            ):  # Skip header row (start from 1-based index)
                row = sheet[row_num]
                if ExcelUtils.is_row_empty(row):
                    continue

                # Get content from columns 1-6 (0-based index 0-5)
                serial_no_cell = row[0]
                analysis_model_id_cell = row[1]
                question_cell = row[2]
                self_define_tags_cell = row[3]
                knowledge_cell = row[4]
                llm_output_cell = row[5]
                prompt_cell = row[8]

                # Build input model
                input_model = BaseInputModel(
                    serial_no=int(
                        ExcelUtils.get_cell_value_as_string(serial_no_cell) or "0"
                    ),
                    analysis_model_id=ExcelUtils.get_cell_value_as_string(
                        analysis_model_id_cell
                    ),
                    question=ExcelUtils.get_cell_value_as_string(question_cell),
                    self_define_tags=ExcelUtils.get_cell_value_as_string(
                        self_define_tags_cell
                    ),
                    llm_output=ExcelUtils.get_cell_value_as_string(llm_output_cell),
                    knowledge=ExcelUtils.get_cell_value_as_string(knowledge_cell),
                    prompt=ExcelUtils.get_cell_value_as_string(prompt_cell),
                )

                input_list.append(input_model)

            input_sets.data_list = input_list
        except Exception as e:
            logger.error(f"parse excel error, path: {path}, errorMsg: {e}")
        finally:
            try:
                if workbook is not None:
                    workbook.close()
            except Exception as e:
                logger.error(f"close workbook error, path: {path}, errorMsg: {e}")

        return input_sets

    def parse_standard_benchmark_sets(
        self, standard_excel_path: str, evaluation_env: EvaluationEnv = EvaluationEnv.DEV
    ) -> List[AnswerExecuteModel]:
        df = pd.read_excel(standard_excel_path, sheet_name=0)
        outputs: List[AnswerExecuteModel] = []
        for _, row in df.iterrows():
            try:
                serial_no = int(row["编号"])
            except Exception:
                continue
            question = row.get("用户问题")
            analysis_model_id = row.get("数据集ID")
            llm_output = (
                None if pd.isna(row.get("标准答案SQL")) else str(row.get("标准答案SQL"))
            )
            order_by = True
            if not pd.isna(row.get("是否排序")):
                try:
                    order_by = bool(int(row.get("是否排序")))
                except Exception:
                    order_by = True

            std_result: Optional[List[Dict[str, List[str]]]] = None
            if not pd.isna(row.get("标准结果")):
                std_result_raw = str(row.get("标准结果"))
                std_result = self._parse_multi_standard_result(std_result_raw)

            strategy_config = DataCompareStrategyConfig(
                strategy="CONTAIN_MATCH",
                order_by=order_by,
                standard_result=std_result if std_result is not None else None,
            )
            outputs.append(
                AnswerExecuteModel(
                    serialNo=serial_no,
                    analysisModelId=analysis_model_id,
                    question=question,
                    llmOutput=llm_output,
                    executeResult=std_result,
                    strategyConfig=strategy_config,
                )
            )
        return outputs

    def _parse_multi_standard_result(
        self, std_result_raw: str
    ) -> Optional[List[Dict[str, List[str]]]]:
        """
        Parse multiple standard results from raw string data.

        Handles multiple results separated by newlines and parses each line as a dict.

        Args:
            std_result_raw (str): Raw standard result string with multiple lines

        Returns:
            Optional[List[Dict[str, List[str]]]]: List of parsed dictionaries,
            or None if parsing fails or no valid data
        """
        try:
            std_result_raw = std_result_raw.strip()
            if not std_result_raw:
                return None

            # 处理多个结果，通过换行符分隔
            result_lines = std_result_raw.split("\n")
            result_list = []

            for line in result_lines:
                line = line.strip()
                if line:
                    try:
                        result_list.append(json.loads(line))
                    except Exception as e:
                        logger.warning(
                            f"Failed to parse line as JSON: {line}, error: {e}"
                        )
                        continue

            return result_list if result_list else None
        except Exception as e:
            logger.error(f"parse multiple standard results error: {e}")
            return None
