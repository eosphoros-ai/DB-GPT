"""Excel Knowledge."""

import logging
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd

from dbgpt.core import Document
from dbgpt.rag.knowledge.base import (
    ChunkStrategy,
    DocumentType,
    Knowledge,
    KnowledgeType,
)


class ExcelKnowledge(Knowledge):
    """Excel Knowledge."""

    def __init__(
        self,
        file_path: Optional[str] = None,
        knowledge_type: Optional[KnowledgeType] = KnowledgeType.DOCUMENT,
        source_columns: Optional[str] = None,
        encoding: Optional[str] = "utf-8",
        loader: Optional[Any] = None,
        metadata: Optional[Dict[str, Union[str, List[str]]]] = None,
        **kwargs: Any,
    ) -> None:
        """Create xlsx Knowledge with Knowledge arguments.

        Args:
            file_path(str,  optional): file path
            knowledge_type(KnowledgeType, optional): knowledge type
            source_column(str, optional): source column
            encoding(str, optional): csv encoding
            loader(Any, optional): loader
        """
        super().__init__(
            path=file_path,
            knowledge_type=knowledge_type,
            data_loader=loader,
            metadata=metadata,
            **kwargs,
        )
        self._encoding = encoding
        self._source_columns = source_columns.split(",") if source_columns else None

    def _find_header_row(
        self, data: List[List[Any]], max_rows_to_check: int = 5
    ) -> Tuple[List[str], int]:
        """
        find header row.
        """
        if not data or not data[0]:
            return [], -1

        num_cols = len(data[0])
        best_headers = []
        best_row_index = -1
        highest_score = -1

        for r_idx in range(min(len(data), max_rows_to_check)):
            row_data = data[r_idx]

            if (
                sum(1 for x in row_data if x is not None and str(x).strip() != "")
                < num_cols / 4
            ):
                continue

            potential_headers = [
                str(x).strip() if x is not None else "" for x in row_data
            ]

            cleaned_headers = [
                h
                for h in potential_headers
                if h and not h.replace(".", "", 1).isdigit()
            ]

            score = len(cleaned_headers) * 2 + len(set(cleaned_headers))

            if not any(h for h in cleaned_headers):
                continue

            avg_len = sum(len(h) for h in cleaned_headers) / (
                len(cleaned_headers) if cleaned_headers else 1
            )
            if avg_len < 2 and len(cleaned_headers) < num_cols / 2:
                continue

            if score > highest_score:
                highest_score = score
                best_headers = potential_headers
                best_row_index = r_idx

        if not best_headers:
            logging.warning(
                f"No clear header row found in the first {max_rows_to_check} rows. "
                f"Using default numeric columns."
            )
            best_headers = [str(i) for i in range(num_cols)]
            best_row_index = -1

        final_headers = []
        header_counts = {}
        for h in best_headers:
            original_h = h
            if h in header_counts:
                header_counts[h] += 1
                h = f"{original_h} ({header_counts[original_h]})"
            else:
                header_counts[h] = 0
            final_headers.append(h if h else f"Col_{len(final_headers)}")

        return final_headers, best_row_index

    def _load(self) -> List[Document]:
        """Load excel document,
        handling merged cells and intelligently identifying headers."""
        docs = []
        if not self._path:
            raise ValueError("file path is required")

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(self._path)
        except Exception as e:
            raise IOError(f"Could not load Excel file with openpyxl: {e}")

        base_metadata = {
            "source": self._path,
            "data_type": "excel",
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            max_row = sheet.max_row
            max_col = sheet.max_column

            if max_row == 0 or max_col == 0:
                logging.info(f"Sheet '{sheet_name}' is empty, skipping.")
                continue

            unmerged_data = [[None for _ in range(max_col)] for _ in range(max_row)]

            for r_idx, row_cells in enumerate(
                sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
            ):
                for c_idx, cell in enumerate(row_cells):
                    unmerged_data[r_idx][c_idx] = cell.value

            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col_merged, max_row_merged = merged_range.bounds

                merged_value = sheet.cell(row=min_row, column=min_col).value

                for r in range(min_row, max_row_merged + 1):
                    for c in range(min_col, max_col_merged + 1):
                        if r - 1 < len(unmerged_data) and c - 1 < len(unmerged_data[0]):
                            unmerged_data[r - 1][c - 1] = merged_value

            headers, header_row_idx = self._find_header_row(unmerged_data)

            if header_row_idx != -1:
                data_rows = unmerged_data[header_row_idx + 1 :]
                df_start_row_index = header_row_idx + 1
                logging.info(
                    f"Sheet '{sheet_name}': Found header row at Excel "
                    f"row {header_row_idx + 1}. Data starts from Excel "
                    f"row {df_start_row_index + 1}."
                )
            else:
                data_rows = unmerged_data
                logging.info(
                    f"Sheet '{sheet_name}': No clear header row found. "
                    f"All rows considered data."
                )

            if not data_rows:
                logging.info(
                    f"Sheet '{sheet_name}': No data rows found after header "
                    f"detection, skipping."
                )
                continue

            processed_data_rows = []
            for row in data_rows:
                if len(row) < max_col:
                    processed_data_rows.append(row + [None] * (max_col - len(row)))
                elif len(row) > max_col:
                    processed_data_rows.append(row[:max_col])
                else:
                    processed_data_rows.append(row)

            df = pd.DataFrame(processed_data_rows)
            df.columns = headers

            df.dropna(axis=1, how="all", inplace=True)

            df = df.loc[:, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)]

            final_headers = df.columns.tolist()
            if not final_headers:
                logging.warning(
                    f"Sheet '{sheet_name}': No valid columns remaining after "
                    f"cleanup, skipping."
                )
                continue

            for index, row in df.iterrows():
                current_metadata = base_metadata.copy()
                current_metadata["sheet_name"] = sheet_name
                current_metadata["row"] = index + df_start_row_index + 1

                strs = []

                for header_name in final_headers:
                    value = row.get(header_name)

                    if header_name is None or pd.isna(value):
                        continue

                    header_str = str(header_name).strip()
                    value_str = str(value).strip() if not pd.isna(value) else ""

                    if header_str:
                        current_metadata[header_str] = value_str

                    if self._source_columns:
                        if header_str in self._source_columns:
                            processed_value = self.parse_document_body(value_str)
                            strs.append(f"{header_str}: {processed_value}")
                    else:
                        processed_value = self.parse_document_body(value_str)
                        strs.append(f"{header_str}: {processed_value}")

                content = "\n".join(strs)

                if not content.strip():
                    logging.debug(
                        f"Skipping empty content document for "
                        f"row {current_metadata['row']} in sheet {sheet_name}"
                    )
                    continue

                if self._metadata:
                    current_metadata.update(self._metadata)

                import uuid

                if "doc_id" not in current_metadata:
                    current_metadata["doc_id"] = str(uuid.uuid4())

                doc = Document(
                    content=content,
                    metadata=current_metadata,
                )
                docs.append(doc)

            return docs

    @classmethod
    def support_chunk_strategy(cls) -> List[ChunkStrategy]:
        """Return support chunk strategy."""
        return [
            ChunkStrategy.CHUNK_BY_SIZE,
            ChunkStrategy.CHUNK_BY_SEPARATOR,
        ]

    @classmethod
    def default_chunk_strategy(cls) -> ChunkStrategy:
        """Return default chunk strategy."""
        return ChunkStrategy.CHUNK_BY_SIZE

    @classmethod
    def type(cls) -> KnowledgeType:
        """Knowledge type of CSV."""
        return KnowledgeType.DOCUMENT

    @classmethod
    def document_type(cls) -> DocumentType:
        """Return document type."""
        return DocumentType.EXCEL

    @property
    def suffix(self) -> Any:
        """Get document suffix."""
        return DocumentType.EXCEL.value
